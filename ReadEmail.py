#import packages
import os
import mailbox
import email
from email import policy
from email.parser import BytesParser
import openpyxl
import datetime
import time
from email.utils import parsedate_to_datetime
from datetime import datetime, timedelta
from email import header
from openpyxl.workbook import Workbook
import schedule

def generate_unique_filename(output_dir,filename):
    """Generate an unique filename by appending a counter or timestamp"""
    base,extension = os.path.splitext(filename)
    counter=1
    unique_filename = filename

    """Check if the file already exists and generate a new name"""
    while os.path.exists(os.path.join(output_dir,unique_filename)):
        unique_filename = f"{base}_{counter}{extension}"
        counter+=1
    return unique_filename

def safe_windows_path(filepath):
    """Rename file path with limitation characters"""
    max_path = 250 #for safely running
    base_dir = os.path.abspath(inbox_dir)
    ext = os.path.splitext(filepath)[1]
    filename = os.path.split(filepath)[1]

    """Count max character for filename"""
    max_filename_len = max_path - len(base_dir) - len(os.sep) - len(ext)

    # Minimal len filename
    max_filename_len = max(10, max_filename_len)
    name_only = os.path.splitext(filename)[0]
    safe_name = name_only[:max_filename_len]
    return os.path.join(base_dir, safe_name + ext)

def save_attachment(part, filename, email_from,subject,date):
    """Save the attachment if it has not been saved before"""
    # create a file path to save the attachment
    unique_filename = generate_unique_filename(inbox_dir,filename)
    #save the attachment
    filepath = os.path.join(inbox_dir,unique_filename)
    filepath = safe_windows_path(filepath)
    with open(filepath, 'wb') as f:
        f.write(part)
        print(f"Attachment saved: {filepath} from Subject: {subject}")

    #update the excel tracking file
    update_tracking_file(unique_filename,email_from,subject,date)
    return True

def update_tracking_file(filename, email_from, subject,date):
    """Update the Excel tracking file with the latest attachment info"""
    if os.path.exists(excel_tracking_file):
        workbook = openpyxl.load_workbook(excel_tracking_file)
        sheet = workbook["Tracking Email"]
    else:
        workbook = Workbook()
        workbook.active.title = "Tracking Email"
        #add headers for the first time
        sheet = workbook.active
        sheet.append(["Run Time", "Subject", "Sender", "Email Received Time", "Filename Attachment"])

    # append the latest data
    tracking_data = (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), subject, email_from, date, filename)
    sheet.append(tracking_data)
    workbook.save(excel_tracking_file)
    print(f"Tracking updated in {excel_tracking_file}")

def load_processed_subjects_from_excel():
    """Generate an Excel tracking file to track and record reading email process"""
    processed=set()
    try:
        workbook = openpyxl.load_workbook(excel_tracking_file)
        sheet = workbook["Tracking Email"]
        for row in sheet.iter_rows(min_row=2,values_only=True):
            subject = row[1]
            received_date_str = row[3]
            if subject and received_date_str:
                processed.add((subject,received_date_str))
    except FileNotFoundError:
        workbook = Workbook()
        # add headers for the first time
        sheet = workbook.create_sheet(title="Tracking Email",index=0)
        sheet.append(["Run Time", "Subject", "Sender", "Email Received Time", "Filename Attachment"])
        workbook.save(excel_tracking_file)
        print(f"Tracking updated in {excel_tracking_file}")
    return processed

def encode_subject(subject):
    """Cleaning Encoded Subject in utf-8 format to convert Chinese Characters"""
    decode_header = header.decode_header(subject)
    subject_clean = ""
    for part, encoding in decode_header:
        if isinstance(part,bytes):
            subject_clean += part.decode(encoding or 'utf-8',errors='replace')
        else:
            subject_clean += part
    return subject_clean


def check_emails(mbox_path,inbox_dir,excel_tracking_file,day_yesterday):
    print(f"Checking emails at {datetime.now().strftime('%H:%M')}")

    """Initialize Inbox Folder to Save Attachment Email"""
    if not os.path.exists(inbox_dir):
        os.makedirs(inbox_dir)

    # get today's date in the required format
    today = datetime.now().date()
    yesterday = today - timedelta(days=day_yesterday) # define range time to read Emails

    #load previously processed subjects from excel
    processed_subjects = load_processed_subjects_from_excel()

    # open the mbox file
    mbox = mailbox.mbox(mbox_path)

    #loop through all the emails in the mbox file
    for message in reversed(mbox):
        email_date_str = message['Date']
        #decode email date and convert it to a datetime object
        email_date = email.utils.parsedate_to_datetime(email_date_str).date()
        received_date = email.utils.parsedate_to_datetime(email_date_str).strftime('%Y-%m-%d %H:%M:%S')

        if email_date > yesterday:
            # get the email sender
            email_from = encode_subject(message['From'])
            subject = encode_subject(message['Subject'])

            if (subject,received_date) in processed_subjects:
                print(f"Skipping email with subject: {subject} {received_date} already processed")
                continue

            msg = BytesParser(policy=policy.default).parsebytes(message.as_bytes())

            if msg.is_multipart():
                # loop through the email parts to find attachments
                for part in msg.walk():
                    if part.get_content_disposition() == 'attachment':
                        filename = part.get_filename()
                        if filename:
                            save_attachment(part, filename, email_from, subject, received_date)
                        else:
                            print(f"CHECK MANUAL from Subject: {subject} - {received_date}")
                            update_tracking_file(filename, email_from, subject, received_date)
                    else:
                        print(f"There is no Attachment from Subject: {subject} - {received_date}")
                        update_tracking_file("", email_from, subject, received_date)
        elif email_date < yesterday: break

def run_program_all(scheduled_time):
    print(f"Saving emails at {datetime.now().strftime('%H:%M')} (scheduled for: {scheduled_time})")
    check_emails(mbox_path, inbox_dir, excel_tracking_file, day_yesterday)

def schedule_tasks():
    """Schedule Running Time Daily in Hours"""
    schedulers = ["08:00", "10:00", "13:00", "15:00"]
    for hour in schedulers:
        schedule.every().day.at(hour).do(run_program_all, hour)

def run_pending_task():
    """While the Schedule is not running, program waits for the next schedule"""
    while True:
        schedule.run_pending()
        time.sleep(1)


#start the email check process
if __name__ == '__main__':
    # path to your thunderbird mbox file
    mbox_path = r"C:\Users\<username>\AppData\Roaming\Thunderbird\Profiles\<profile>\Mail"
    # directory to save attachments
    inbox_dir = r"C:\Users\Inbox_Email"
    # path to Excel Tracking File
    excel_tracking_file = r"C:\Users\Thunderbird_Project\Tracking\Tracking_Email.xlsx"
    # define a list of unwanted file extensions
    wanted_extensions = {'.xls', '.xlsx', '.xlsm', '.xlsb', '.ods'}
    day_yesterday = 2
    print("Scheduler started. Waiting for tasks..")
    schedule_tasks()
    run_pending_task()

